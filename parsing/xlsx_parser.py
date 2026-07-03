"""
.xlsx parser — single deterministic method: openpyxl.

Also extracts embedded images via `ws._images` (private-ish but the only
practical API openpyxl exposes for this; confirmed working against 2 real
corpus files with real embedded images before relying on it here) — a.py's
original XLSXParser was missing this.
"""
from pathlib import Path

import openpyxl

from .common import ImageWriter, ParsedBlock, ParsedDocument


def parse(path: Path, image_writer: ImageWriter) -> ParsedDocument:
    wb = openpyxl.load_workbook(path, data_only=True)
    parsed = ParsedDocument(source_file=path.name, doc_type="xlsx", units_total=len(wb.sheetnames))
    parsed.metadata = {"sheets": len(wb.sheetnames)}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        parsed.blocks.append(ParsedBlock(kind="heading", content=f"Лист: {sheet_name}", level=2))

        rows = [[("" if c is None else str(c)) for c in row] for row in ws.iter_rows(values_only=True)]
        if rows:
            parsed.blocks.append(ParsedBlock(kind="table", content=rows))

        for img in getattr(ws, "_images", []):
            try:
                data = img._data()
                ext = Path(img.path).suffix.lstrip(".") or "png"
                image_id = image_writer.add(data, ext)
                parsed.blocks.append(ParsedBlock(kind="image", image_id=image_id))
            except Exception as e:
                parsed.blocks.append(ParsedBlock(kind="text", content=f"[image extraction error: {e}]"))

    return parsed
